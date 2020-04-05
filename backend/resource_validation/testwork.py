import openpyxl as pyxl
from datetime import date, time, timedelta
import re
import sys
import platform
import colorsys
import itertools
import platform
from resource_validation.excel_funcs import excel_funcs

def generateCalendar(startDate, numWeeks):
	startToEnd = []
	dayOfWk = startDate.weekday()
	for i in range(dayOfWk, 0, -1):
		startDate -= timedelta(days = 1)
	curDate = startDate
	for i in range(0, numWeeks):
		for i in range(0, 7):
			startToEnd.append(curDate)
			curDate += timedelta(days = 1)
	return startToEnd

def getNames(st):
	names = {}
	nameRange = st.range('C:C').api.Find(' ')
	nameRange = st.range('C:C').api.FindNext(nameRange)
	while nameRange != None:
		if str(nameRange.value) == 'end line':
			break
		lname = str(nameRange.value).split(' ')
		fname = lname[1]
		lname = lname[0]
		names[lname] = fname
		nameRange = st.range('C:C').api.FindNext(nameRange)

##########################################################

def resource_validation(file_path):
	wb = pyxl.load_workbook(file_path)
	st = wb[wb.sheetnames[0]]
	if "Calendar" in wb.sheetnames:
		wb.remove(wb["Calendar"])
	ws = wb.create_sheet("Calendar")

	wb.save(file_path)


	# ##########################################################

	# nameList = getNames(st)

	# #########################################################
	# rangeRow = st.range('A1').api.EntireRow(4)
	# weeksRow = rangeRow.Find('Sum of')
	# firstWeek = str(weeksRow.value)
	# numWeeks = 0
	# while weeksRow != None:
	# 	if numWeeks != 0 and firstWeek == str(weeksRow.value):
	# 		break
	# 	numWeeks += 1
	# 	weeksRow = rangeRow.FindNext(weeksRow)

	# dates = generateCalendar(startDate, numWeeks)

	# calendarRange = st.range('N4')
	# count = 0
	# for date in dates:
	# 	calendarRange.clear()
	# 	if count == 0:
	# 		calendarRange.color = 0x0095dc
	# 		calendarRange = calendarRange.offset(column_offset=1)
	# 		calendarRange.clear()
		
	# 	calendarRange.value = date.strftime('%m/%d/%Y')
	# 	calendarRange.autofit()
	# 	count = (count + 1) % 7
	# 	calendarRange = calendarRange.offset(column_offset=1)

	# weekDay = st.range('D5').value
	# # for weekNo in range()