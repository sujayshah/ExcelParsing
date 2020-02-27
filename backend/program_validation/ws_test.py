import openpyxl as pyxl
from openpyxl import formula
from openpyxl.styles import Color, PatternFill, Alignment, Font, Border, Side
from openpyxl.cell import Cell
from openpyxl.worksheet import dimensions
from datetime import date, timedelta, datetime
import platform
import colorsys
import pandas as pd

class EXCEL_FUNCS:
	def __init__(self, func_type):
		self.func_type = func_type

	def runFunction(self, st, op_range, cellCache, offset = 0, secondary_range = None): 
		if self.func_type == 'MIN(':
			# print("getMin")
			return self.getMin(st, op_range, cellCache)
		elif self.func_type == 'MAX(':
			return self.getMax(st, op_range, cellCache)
			# print("getMax")
		elif self.func_type == 'WORKDAY(':
			# print("getWorkday")
			return self.getWorkday(st, op_range, cellCache, offset, secondary_range)
		elif op_range:
			# print("getCellValue")
			return self.getCellValue(st, op_range, offset, cellCache)
		else:
			return None

	def getCellValue(self, st, op_range, offset, cellCache):
		cellValue = st[op_range].value
		if not isinstance(cellValue, datetime):
			cellValue = evaluateFormulaToDate(st, cellValue, cellCache)
		return cellValue + timedelta(days = offset)

	def getMin(self, st, op_range, cellCache):
		min = None
		try:
			for row in st[op_range]:
				for cell in row:
					evalCell = cell.value
					if not isinstance(cell.value, datetime):
						evalCell = evaluateFormulaToDate(st, cell.value, cellCache)
					if not min:
						min = evalCell
					elif evalCell < min:
						min = evalCell
		except Exception as e:
			print(e)
		return min

	def getMax(self, st, op_range, cellCache):
		max = None
		try:
			for row in st[op_range]:
				for cell in row:
					evalCell = cell.value
					if not isinstance(cell.value, datetime):
						evalCell = evaluateFormulaToDate(st, cell.value, cellCache)
					if not max:
						max = evalCell
					elif evalCell > max:
						max = evalCell
		except Exception as e:
			print(e)
		return max
	
	def getWorkday(self, st, op_range, cellCache, offset, secondary_range):
		workday = st[op_range].value
		try:
			if not isinstance(workday, datetime):
				workday = evaluateFormulaToDate(st, workday, cellCache)
			if workday.weekday() >= 5:
				workday -= timedelta(days = workday.weekday() - 4)
			for i in range(0, offset):
				if workday.weekday() >= 5:
					workday += timedelta(days = 7 - workday.weekday())
				workday += timedelta(days = 1)
		except Exception as e:
			print(e)
		while workday.weekday() >= 5:
			workday += timedelta(days = 1)
		return workday

def generateCalendar(startDate, endDate):
	startToEnd = []
	curDate = startDate
	while curDate <= endDate:
		startToEnd.append(curDate)
		# startToEnd.append(curDate.strftime("X%m/X%d/%Y").replace('X0','X').replace('X',''))
		curDate += timedelta(days = 7)
	return startToEnd

def evaluateFormulaToDate(st, evalString, cellCache):
	if isinstance(evalString, datetime):
		return evalString
	else:
		tok = formula.Tokenizer(evalString)
		func_type = next((t for t in tok.items if t.type == 'FUNC' and t.subtype == 'OPEN'), None)
		operand_range_iter = (t for t in tok.items if t.type == 'OPERAND' and t.subtype == 'RANGE')
		operand_range = next(operand_range_iter, None)
		secondary_range = next(operand_range_iter, None)
		offset = next((t for t in tok.items if t.type == 'OPERAND' and t.subtype == 'NUMBER'), 0)
		offset_infix = next((t for t in tok.items if t.type == 'OPERATOR-INFIX'), 0)
		# print(func_type, operand_range, secondary_range, offset)
		# print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))

		if func_type:
			func_type = func_type.value
		formula_func = EXCEL_FUNCS(func_type)

		if operand_range:
			operand_range = operand_range.value

		if secondary_range:
			secondary_range = secondary_range.value

		if isinstance(offset, pyxl.formula.tokenizer.Token) and isinstance(offset.value, str):
			offset = int(float(offset.value))
			if offset_infix and offset_infix.value == '-':
				offset *= -1

		funcOutput = formula_func.runFunction(st=st, op_range=operand_range, cellCache=cellCache, 
		offset=offset, secondary_range=secondary_range)
		return funcOutput

def parseSchedule(st, userStartDate, userEndDate):
	categories = []
	eventList = []
	cellCache = {}
	class Event():
		def __init__(self, activity, start, finish): 
			self.machineID = activity[activity.find('(')+1:activity.find(')')].strip()
			self.activityID = activity[:activity.find('(')].strip()
			self.startDate = start
			self.finishDate = finish
			self.subTasks = []
	class Task():
		def __init__(self, category, task, start, finish):
			self.category = category
			self.task = task
			self.startDate = start
			self.finishDate = finish

	# output = evaluateFormulaToDate(st, st['D2'].value, cellCache)
	# print(output)

	########################################
	for a, b, c, d in zip(st['A'], st['B'], st['C'], st['D']):
		# Title Row
		if a.row == 1:
			continue
		# Check if Date
		startDate = evaluateFormulaToDate(st, c.value, cellCache)
		endDate = evaluateFormulaToDate(st, d.value, cellCache)
		# print(startDate, endDate)

		if a.value == None:
			eventList.append(Event(b.value, startDate, endDate))
		if a.value != None:
			eventList[-1].subTasks.append(Task(a.value, b.value, startDate, endDate))
			if a.value not in categories:
				categories.append(a.value)

	return categories, eventList

def styleCell(curCell, fillColor, autofit = False, mergedCells = True):
	curCell.fill = PatternFill(fill_type='solid', start_color=fillColor, end_color=fillColor)	
	curCell.font = Font(name='Arial', size=14)
	curCell.alignment = Alignment(horizontal = 'center', vertical='center')

	curSide = Side(color='00000000', border_style='thick')
	curCell.border = Border(top = curSide, bottom = curSide, left = curSide, right = curSide)
	if mergedCells:
		for i in range(1, 7):
			curCell.offset(column = i).border = Border(top = curSide, bottom = curSide)
	if autofit:
		autoLength = len(curCell.value)*1.8
		col_dim = curCell.parent.column_dimensions[curCell.column_letter]
		if not col_dim.min:
			col_dim.min = autoLength
		elif autoLength > col_dim.min:
			col_dim.min = autoLength
		curCell.parent.column_dimensions[curCell.column_letter].width = col_dim.min


def writeExcel(st, wkList, eventData, palette, categories, startDate, endDate):
	st['A1'] = 'Machine S/N'
	st['B1'] = 'Phase'
	styleCell(st['A1'], "FFFFFFFF", autofit = True, mergedCells=False)
	styleCell(st['B1'], "FFFFFFFF", autofit = True, mergedCells=False)

	curCell = st['C1']
	st.row_dimensions[1] = dimensions.RowDimension(worksheet = st, height = 50)
	for idx, wkCell in enumerate(wkList):
		curCell.value = wkList[idx].strftime("X%m/X%d/%Y").replace('X0','X').replace('X','')
		st.merge_cells(start_column = curCell.column, end_column = curCell.column + 6, start_row = curCell.row, end_row = curCell.row)		
		styleCell(curCell, "FFFFA500")
		curCell = curCell.offset(column = 7)

	machineCol = st['A2']
	activityCol = st['B2']
	calendarStartCol = st['C2']
	calendarEndCol = calendarStartCol.offset(column = len(wkList)*7-1)

	for event in iter(eventData):
		machineCol.value = event.machineID
		activityCol.value = event.activityID
		styleCell(machineCol, "00FFFF00", autofit=True, mergedCells=False)
		styleCell(activityCol, "00FFFF00", autofit=True, mergedCells=False)
		
		for task in event.subTasks:
			taskStart = task.startDate
			taskEnd = task.finishDate
			if isinstance(taskStart, datetime):
				taskStart = taskStart.date()
			if isinstance(taskEnd, datetime):
				taskEnd = taskEnd.date()
			# if taskStart < startDate:
			# 	raise IndexError("Task " + task.task + " has an starting date before your specified start date. Please choose an earlier start date.")
			# else:
			taskLength = (taskEnd - taskStart).days
			try:
				offsetCell = calendarStartCol.offset(column = (taskStart - startDate).days)
				offsetCell.value = task.task
				if taskLength > 0:
					st.merge_cells(start_column = offsetCell.column, end_column = offsetCell.column + taskLength, start_row = offsetCell.row, end_row = offsetCell.row)
			except:
				raise AttributeError("Task " + task.task + " is overlapping with another task in activity " + activityCol.value)

		machineCol = machineCol.offset(row = 1)
		activityCol = activityCol.offset(row = 1)
		calendarStartCol = calendarStartCol.offset(row = 1)
		calendarEndCol = calendarEndCol.offset(row = 1)

#--------------------------------------------------------

def program_validation(file_path, palette, start, end):
	wkList = generateCalendar(start, end)
	wb = pyxl.load_workbook(file_path)
	st = wb[wb.sheetnames[0]]
	try:
		categories, eventList = parseSchedule(st, start, end)
		if "Calendar" in wb.sheetnames:
			wb.remove(wb["Calendar"])
		ws = wb.create_sheet("Calendar")
		writeExcel(ws, wkList, eventList, palette, categories, start, end)
		# writeExcel(ws, wkList, eventList, palette, categories, start.strftime("X%m/X%d/%Y").replace('X0','X').replace('X',''))
		wb.save(file_path)
	except AttributeError as ae:
		raise ae
	except IndexError as ie:
		raise ie
	except ValueError as ve:
		raise Exception('Invalid date or offset')
	except TypeError as te:
		raise Exception('A cell in column C or D has an invalid date or formula')

if __name__ == "__main__":
	palette = ['#ADD8E6', ' #90EE90', '#FFB6C1',  '#DDBDF1','#FFD700','#FFDAB9','#FF69B4','#7FFFD4', '#DEB887', '#C0C0C0' ]
	program_validation('./backend/program_validation/input/sample.xlsx', palette, date(year=2019, month=8, day=1), date(year=2022, month=3, day=1))